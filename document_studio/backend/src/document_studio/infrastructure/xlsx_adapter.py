"""XLSX ingestion adapter for Document Studio.

Uses ``openpyxl`` to read workbook structure directly, preserving sheet
names, cell coordinates, and useful cell values.  Spreadsheets are
never rendered as images.

Parser imports are kept inside this infrastructure module.
"""

from __future__ import annotations

from document_studio.domain.normalized import (
    CellLocation,
    ContentBlock,
    IngestionOutcome,
    IngestionResult,
    NormalizedDocument,
    NormalizedSheet,
    TableBlock,
    TableCell,
)


def _col_letter(col_idx: int) -> str:
    """Convert a 1-indexed column number to an Excel column letter.

    For example: 1 -> 'A', 26 -> 'Z', 27 -> 'AA'.
    """
    result = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def ingest_xlsx(data: bytes, original_filename: str) -> IngestionResult:
    """Ingest an XLSX workbook from raw bytes.

    Returns a successful ``IngestionResult`` with sheet names,
    cell coordinates, and cell values preserved.  Each worksheet
    becomes a ``NormalizedSheet`` containing a single ``TableBlock``
    with all non-empty cells.
    """
    import io

    import openpyxl  # imported here to keep parser deps in infra

    wb = openpyxl.load_workbook(
        io.BytesIO(data), read_only=True, data_only=True
    )

    try:
        sheets: list[NormalizedSheet] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            cells: list[TableCell] = []
            max_row = 0
            max_col = 0

            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        if text:
                            cells.append(
                                TableCell(row=row_idx, col=col_idx, text=text)
                            )
                            if row_idx + 1 > max_row:
                                max_row = row_idx + 1
                            if col_idx + 1 > max_col:
                                max_col = col_idx + 1

            # Build a table block for this sheet.
            # Each cell also gets a cell_ref for source-location clarity.
            table_block = TableBlock(
                cells=tuple(cells),
                row_count=max_row,
                col_count=max_col,
                location=CellLocation(sheet_name=sheet_name),
            )

            blocks: list[ContentBlock] = [table_block] if cells else []

            sheets.append(
                NormalizedSheet(
                    name=sheet_name,
                    blocks=tuple(blocks),
                )
            )

        # Filter out completely empty sheets to satisfy the contract
        # that at least one meaningful sheet exists.
        non_empty = [s for s in sheets if s.blocks]
        if not non_empty:
            # If every sheet is empty, still include all sheets with
            # empty blocks so the workbook structure is visible.
            # But we need at least one sheet to satisfy NormalizedDocument.
            non_empty = sheets if sheets else []

        # If the workbook has no sheets at all, this is an edge case;
        # return the sheet list as-is (NormalizedDocument will validate).
        return IngestionResult(
            outcome=IngestionOutcome.success,
            document=NormalizedDocument(
                media_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                original_filename=original_filename,
                sheets=tuple(non_empty if non_empty else sheets),
            ),
        )

    finally:
        wb.close()
